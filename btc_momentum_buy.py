#!/usr/bin/env python3
"""
BTC Momentum Scoring System v3.0 — BUY Side
JP Trust Learning

Adapted from Gold Momentum Scoring v3.0
- All dimensions D1-D6 scored 0-100 each
- Total score = average of all 6 dimensions (0-100)
- Penalty scaled proportionally
- Net Score = Gross (avg D1-D6) + Penalty_Scaled

Key differences from Gold:
- D5 Volatility thresholds adjusted for crypto (BTC typically 40-80% annualized)
- D6 External: uses DXY + VIX but with crypto-specific interpretation
  (BTC = risk-on asset, opposite of gold's safe-haven behavior)
- No weekend filtering needed (BTC trades 24/7)
- Volume in BTC units (not contracts)
"""

import pandas as pd
import numpy as np
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys

# ── CONFIG ──
ROLLING_WINDOW = 252
LOOKBACK = {'1W': 5, '1M': 21, '3M': 63, '6M': 126, '1Y': 252}
WEIGHTS = {'1Y': 0.30, '6M': 0.25, '3M': 0.20, '1M': 0.15, '1W': 0.10}
WEIGHT_ORDER = ['1Y', '6M', '3M', '1M', '1W']

RUN_TS = datetime.now(timezone.utc)
AS_OF = RUN_TS.strftime("%d/%m/%Y %H:%M UTC")
TS_FILE = RUN_TS.strftime("%d%m%Y_%H%M")

# ── LOAD DATA ──
base_dir = os.path.dirname(os.path.abspath(__file__))

def load_price_csv(filename):
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        print(f"⚠️ {filename} not found — skipping")
        return None
    df = pd.read_csv(path, encoding='utf-8-sig')
    df.columns = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume']
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)
    return df

df = load_price_csv('btc_prices.csv')
df_dxy = load_price_csv('dxy_prices.csv')
df_vix = load_price_csv('vix_prices.csv')

if df is None:
    print("❌ btc_prices.csv not found — cannot continue")
    sys.exit(1)

# ── BASE DATES ──
BD2_idx = len(df) - 1
BD1_idx = len(df) - 6
BD1_date = df.iloc[BD1_idx]['Date']
BD2_date = df.iloc[BD2_idx]['Date']

print(f"BTC Momentum Scoring v3.0 — BUY Side")
print(f"{'='*55}")
print(f"Base Date 1: {BD1_date.strftime('%Y-%m-%d')} (idx={BD1_idx})")
print(f"Base Date 2: {BD2_date.strftime('%Y-%m-%d')} (idx={BD2_idx})")
print(f"Total BTC rows: {len(df)}")
if df_dxy is not None:
    print(f"DXY rows: {len(df_dxy)} (latest: {df_dxy['Date'].max().strftime('%Y-%m-%d')})")
if df_vix is not None:
    print(f"VIX rows: {len(df_vix)} (latest: {df_vix['Date'].max().strftime('%Y-%m-%d')})")


# ══════════════════════════════════════════════════════
# CORE CALCULATION FUNCTIONS
# ══════════════════════════════════════════════════════

def compute_return(closes, end_idx, period_days):
    start_idx = end_idx - period_days
    if start_idx < 0:
        return None
    return (closes[end_idx] - closes[start_idx]) / closes[start_idx] * 100

def rolling_percentile(series_values, current_val, window=ROLLING_WINDOW):
    valid = series_values[~np.isnan(series_values)]
    if len(valid) < 10:
        return 50.0
    count_below = np.sum(valid < current_val)
    return count_below / (len(valid) - 1) * 100 if len(valid) > 1 else 50.0

def calc_return_percentiles(df, base_idx):
    closes = df['Close'].values
    results = {}
    for period, days in LOOKBACK.items():
        current_ret = compute_return(closes, base_idx, days)
        if current_ret is None:
            results[period] = {'return': 0, 'percentile': 50}
            continue
        rolling_rets = []
        start = max(0, base_idx - ROLLING_WINDOW)
        for i in range(start, base_idx):
            r = compute_return(closes, i, days)
            if r is not None:
                rolling_rets.append(r)
        pctl = rolling_percentile(np.array(rolling_rets), current_ret) if rolling_rets else 50
        results[period] = {'return': current_ret, 'percentile': pctl}
    return results

def calc_volume_percentiles(df, base_idx):
    volumes = df['Volume'].values
    results = {}
    for period, days in LOOKBACK.items():
        end = base_idx + 1
        start = end - days
        if start < 0:
            results[period] = {'volume': 0, 'percentile': 50}
            continue
        current_vol = np.sum(volumes[start:end])
        rolling_vols = []
        roll_start = max(0, base_idx - ROLLING_WINDOW)
        for i in range(roll_start, base_idx):
            s = i + 1 - days
            if s < 0:
                continue
            rolling_vols.append(np.sum(volumes[s:i+1]))
        pctl = rolling_percentile(np.array(rolling_vols), current_vol) if rolling_vols else 50
        results[period] = {'volume': current_vol, 'percentile': pctl}
    return results

def weighted_percentile(pctl_dict):
    return sum(pctl_dict[p]['percentile'] * WEIGHTS[p] for p in WEIGHT_ORDER)

def d1_score(wp): return wp  # WP is already 0-100
def d2_score(wp): return wp  # WP is already 0-100

def calc_rsi(df, base_idx, period=14):
    start = base_idx - 29
    if start < 0: start = 0
    closes = df['Close'].values[start:base_idx+1]
    if len(closes) < 2:
        return 50
    deltas = np.diff(closes)
    gains = np.where(deltas > 0, deltas, 0)
    losses = np.where(deltas < 0, -deltas, 0)
    last_n = min(period, len(gains))
    avg_gain = np.mean(gains[-last_n:])
    avg_loss = np.mean(losses[-last_n:])
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def d3_score(rsi):
    """RSI scoring — same logic as gold (sweet spot 50-70)"""
    if 50 <= rsi <= 70: return 100
    if 40 <= rsi < 50: return 80
    if 70 < rsi <= 80: return 70
    if 30 <= rsi < 40: return 60
    if rsi > 80: return 50
    return 30

def calc_ma(df, base_idx, window):
    start = base_idx + 1 - window
    if start < 0: return None
    return np.mean(df['Close'].values[start:base_idx+1])

def d4_score(price, ma50, ma200):
    """MA Trend — same logic as gold"""
    pts = 0
    if ma50 is not None and price > ma50: pts += 35
    if ma200 is not None and price > ma200: pts += 35
    if ma50 is not None and ma200 is not None and ma50 > ma200: pts += 30
    return min(pts, 100)

def calc_volatility(df, base_idx):
    start = base_idx - 20
    if start < 0: start = 0
    closes = df['Close'].values[start:base_idx+1]
    if len(closes) < 2:
        return 0
    rets = np.diff(closes) / closes[:-1]
    return np.std(rets) * np.sqrt(365) * 100  # 365 for crypto (trades every day)

def d5_score(vol):
    """
    D5 Volatility — adjusted for crypto (BTC typically 40-80% annualized).
    Lower vol = more stable uptrend = higher score.
    Thresholds wider than gold because crypto is inherently more volatile.
    """
    if vol <= 30:  return 100
    if vol <= 45:  return 90
    if vol <= 60:  return 70
    if vol <= 75:  return 55
    if vol <= 90:  return 40
    if vol <= 120: return 25
    return 10

def calc_penalties(df, base_idx):
    closes = df['Close'].values
    ret_1y = compute_return(closes, base_idx, 252) or 0
    ret_6m = compute_return(closes, base_idx, 126) or 0
    ret_1m = compute_return(closes, base_idx, 21) or 0
    ret_1w = compute_return(closes, base_idx, 5) or 0

    reversal_pen = 0
    reversal_flag = ""
    strong = (ret_1y > 20 and ret_1m < -5 and ret_1w < -3)
    mild = ((ret_1y > 0 or ret_6m > 0) and ret_1m < 0 and ret_1w < 0)
    if strong:
        reversal_pen = -10
        reversal_flag = "🔴 Strong Reversal"
    elif mild:
        reversal_pen = -5
        reversal_flag = "⚠️ Mild Reversal"

    ma50 = calc_ma(df, base_idx, 50)
    ma200 = calc_ma(df, base_idx, 200)
    price = closes[base_idx]
    dc_pen = 0
    dc_flag = ""
    if ma50 is not None and ma200 is not None and ma50 < ma200:
        dc_pen = -5
        if price < ma50 and price < ma200:
            dc_flag = "💀💀 Death Cross + Below MAs"
        else:
            dc_flag = "💀 Death Cross"

    total = max(reversal_pen + dc_pen, -15)
    flags = " | ".join(f for f in [reversal_flag, dc_flag] if f)
    return {
        'reversal': reversal_pen, 'death_cross': dc_pen,
        'total': total, 'flags': flags,
        'ret_1y': ret_1y, 'ret_6m': ret_6m, 'ret_1m': ret_1m, 'ret_1w': ret_1w
    }


# ══════════════════════════════════════════════════════
# DIMENSION 6 — EXTERNAL CONTEXT (DXY + VIX)
# BTC-specific: BTC is risk-on, OPPOSITE of gold's safe-haven
# ══════════════════════════════════════════════════════

def find_closest_idx(ext_df, target_date, max_gap_days=5):
    if ext_df is None:
        return None
    diffs = (ext_df['Date'] - target_date).abs()
    min_diff = diffs.min()
    if min_diff.days > max_gap_days:
        return None
    return diffs.idxmin()

def calc_external_return(ext_df, end_idx, period_days):
    if ext_df is None or end_idx is None:
        return None
    start_idx = end_idx - period_days
    if start_idx < 0:
        return None
    return (ext_df['Close'].values[end_idx] - ext_df['Close'].values[start_idx]) / ext_df['Close'].values[start_idx] * 100

def calc_d6_external(df_btc, btc_idx, df_dxy, df_vix):
    """
    Dimension 6: External Context Score (±10 pts total)
    
    BTC is RISK-ON asset (opposite of gold):
    
    Part A — DXY (±5 pts):
      BTC up + DXY down (normal risk-on) → +2
      BTC up + DXY up (BTC strong despite strong $) → +5 (divergence = very bullish)
      BTC down + DXY down → 0 (neutral)
      BTC down + DXY up → -5 (risk-off = bearish BTC)
    
    Part B — VIX Regime (±5 pts):
      VIX < 20 + BTC up → +5 (calm market = risk-on = BTC thrives)
      VIX 20-30 + BTC up → +2 (moderate fear but BTC still up)
      VIX > 30 + BTC up → +1 (BTC resilient in panic)
      VIX < 20 + BTC down → 0 (calm but BTC weak)
      VIX 20-30 + BTC down → -3 (fear dragging BTC)
      VIX > 30 + BTC down → -5 (panic selling risk assets)
    
    Total D6 range: -10 to +10
    """
    btc_date = df_btc.iloc[btc_idx]['Date']
    btc_closes = df_btc['Close'].values
    btc_1m = compute_return(btc_closes, btc_idx, 21)
    if btc_1m is None:
        btc_1m = 0
    btc_up = btc_1m >= 0

    # ── Part A: DXY ──
    dxy_score = 0
    dxy_1m = None
    dxy_signal = "N/A"
    
    if df_dxy is not None:
        dxy_idx = find_closest_idx(df_dxy, btc_date)
        if dxy_idx is not None:
            dxy_1m = calc_external_return(df_dxy, dxy_idx, 21)
            if dxy_1m is not None:
                dxy_up = dxy_1m > 0
                if btc_up and dxy_up:
                    dxy_score = +5
                    dxy_signal = "🟢 Bullish Divergence (BTC up despite strong $)"
                elif btc_up and not dxy_up:
                    dxy_score = +2
                    dxy_signal = "🔵 Normal Risk-On (BTC up + weak $)"
                elif not btc_up and not dxy_up:
                    dxy_score = 0
                    dxy_signal = "⚪ Neutral (both down)"
                else:  # btc down, dxy up
                    dxy_score = -5
                    dxy_signal = "🔴 Risk-Off (BTC down + strong $)"

    # ── Part B: VIX Regime ──
    vix_score = 0
    vix_level = None
    vix_signal = "N/A"
    
    if df_vix is not None:
        vix_idx = find_closest_idx(df_vix, btc_date)
        if vix_idx is not None:
            vix_level = df_vix['Close'].values[vix_idx]
            if btc_up:
                if vix_level < 20:
                    vix_score = +5
                    vix_signal = "🟢 Risk-On Rally (VIX<20 + BTC up)"
                elif vix_level <= 30:
                    vix_score = +2
                    vix_signal = "🔵 Moderate Fear (VIX 20-30 + BTC up)"
                else:
                    vix_score = +1
                    vix_signal = "⚪ BTC Resilient (VIX>30 + BTC up)"
            else:
                if vix_level < 20:
                    vix_score = 0
                    vix_signal = "⚪ Calm Drift (VIX<20 + BTC down)"
                elif vix_level <= 30:
                    vix_score = -3
                    vix_signal = "🟠 Fear Dragging BTC (VIX 20-30)"
                else:
                    vix_score = -5
                    vix_signal = "🔴 Panic Sell (VIX>30 + BTC down)"

    total_d6 = max(min(dxy_score + vix_score, 10), -10)
    d6_scaled = (total_d6 + 10) / 20 * 100  # ±10 → 0-100

    return {
        'd6_total': total_d6,
        'd6_scaled': d6_scaled,
        'dxy_score': dxy_score,
        'vix_score': vix_score,
        'dxy_1m': dxy_1m,
        'vix_level': vix_level,
        'dxy_signal': dxy_signal,
        'vix_signal': vix_signal,
        'btc_1m': btc_1m
    }


# ══════════════════════════════════════════════════════
# COMPUTE FULL SCORES
# ══════════════════════════════════════════════════════

def full_score(df, idx, df_dxy, df_vix):
    ret_pctls = calc_return_percentiles(df, idx)
    vol_pctls = calc_volume_percentiles(df, idx)
    wp_ret = weighted_percentile(ret_pctls)
    wp_vol = weighted_percentile(vol_pctls)
    d1 = d1_score(wp_ret)
    d2 = d2_score(wp_vol)
    
    rsi = calc_rsi(df, idx)
    d3 = d3_score(rsi)
    
    price = df['Close'].values[idx]
    ma50 = calc_ma(df, idx, 50)
    ma200 = calc_ma(df, idx, 200)
    d4 = d4_score(price, ma50, ma200)
    
    vol = calc_volatility(df, idx)
    d5 = d5_score(vol)
    
    gross = d1 + d2 + d3 + d4 + d5
    penalties = calc_penalties(df, idx)
    
    ext = calc_d6_external(df, idx, df_dxy, df_vix)
    d6 = ext['d6_scaled']
    
    gross_avg_dims = (d1 + d2 + d3 + d4 + d5 + d6) / 6
    penalty_scaled = penalties['total'] * (100 / 110)
    net = gross_avg_dims + penalty_scaled
    golden_cross = (ma50 is not None and ma200 is not None and ma50 > ma200)
    
    return {
        'date': df.iloc[idx]['Date'],
        'price': price,
        'ret_pctls': ret_pctls, 'vol_pctls': vol_pctls,
        'wp_ret': wp_ret, 'wp_vol': wp_vol,
        'd1': d1, 'd2': d2, 'd3': d3, 'd4': d4, 'd5': d5, 'd6': d6,
        'd6_raw': ext['d6_total'],
        'rsi': rsi, 'ma50': ma50, 'ma200': ma200,
        'golden_cross': golden_cross, 'volatility': vol,
        'gross': gross_avg_dims, 'penalties': penalties,
        'penalty_scaled': penalty_scaled,
        'net': net,
        'external': ext
    }

s1 = full_score(df, BD1_idx, df_dxy, df_vix)
s2 = full_score(df, BD2_idx, df_dxy, df_vix)

net_avg = (s1['net'] + s2['net']) / 2
gross_avg = (s1['gross'] + s2['gross']) / 2
delta = s2['net'] - s1['net']

def tier(score):
    clamped = max(0, min(100, score))
    if clamped >= 85: return "Very Strong ↑↑"
    if clamped >= 75: return "Strong ↑"
    if clamped >= 60: return "Moderate ↑"
    if clamped >= 45: return "Neutral →"
    if clamped >= 30: return "Weak ↓"
    return "Very Weak ↓↓"

momentum_tier = tier(net_avg)

print(f"\n{'='*55}")
print(f"BTC Momentum Score v3.0 — BUY Side")
print(f"{'='*55}")
print(f"Net Score Avg:  {net_avg:.2f}  ({momentum_tier})")
print(f"Gross Score Avg: {gross_avg:.2f}")
print(f"BD1 ({s1['date'].strftime('%Y-%m-%d')}): Net={s1['net']:.2f}  D6={s1['d6']:.1f}/100 (raw {s1['d6_raw']:+d})")
print(f"BD2 ({s2['date'].strftime('%Y-%m-%d')}): Net={s2['net']:.2f}  D6={s2['d6']:.1f}/100 (raw {s2['d6_raw']:+d})")
print(f"Delta: {delta:+.2f}")
print(f"Price: ${s2['price']:.1f}")
print(f"RSI: {s2['rsi']:.1f} | Volatility: {s2['volatility']:.1f}%")
print(f"Penalties: {s2['penalties']['total']} (scaled: {s2['penalty_scaled']:.1f}) ({s2['penalties']['flags'] or 'None'})")
print(f"\n── External Context (BD2) ──")
print(f"D6 Raw: {s2['d6_raw']:+d} → Scaled: {s2['d6']:.1f}/100")
print(f"  DXY: {s2['external']['dxy_score']:+d}  ({s2['external']['dxy_signal']})")
print(f"  VIX: {s2['external']['vix_score']:+d}  ({s2['external']['vix_signal']})")
if s2['external']['dxy_1m'] is not None:
    print(f"  DXY 1M Return: {s2['external']['dxy_1m']:.2f}%")
if s2['external']['vix_level'] is not None:
    print(f"  VIX Level: {s2['external']['vix_level']:.2f}")


# ══════════════════════════════════════════════════════
# MULTI-TF PIVOT POINTS + CONFLUENCE ZONES
# ══════════════════════════════════════════════════════

def calc_pivot_levels(high, low, close):
    PP = (high + low + close) / 3
    R1 = 2 * PP - low
    S1 = 2 * PP - high
    R2 = PP + (high - low)
    S2 = PP - (high - low)
    R3 = high + 2 * (PP - low)
    S3 = low - 2 * (high - PP)
    return {'PP': round(PP, 2), 'R1': round(R1, 2), 'R2': round(R2, 2), 'R3': round(R3, 2),
            'S1': round(S1, 2), 'S2': round(S2, 2), 'S3': round(S3, 2)}

def calc_multi_tf_pivots(df):
    df = df.copy()
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)
    
    result = {}
    current_price = df['Close'].values[-1]
    latest_date = df['Date'].max()
    
    # D1: Last completed day
    if len(df) >= 1:
        prev = df.iloc[-1]
        levels = calc_pivot_levels(prev['High'], prev['Low'], prev['Close'])
        result['D1'] = {**levels, 'H': round(prev['High'], 2), 'L': round(prev['Low'], 2),
                        'C': round(prev['Close'], 2), 'date': prev['Date'].strftime('%Y-%m-%d')}
    
    # W1: Previous completed week
    df['iso_year'] = df['Date'].dt.isocalendar().year.astype(int)
    df['iso_week'] = df['Date'].dt.isocalendar().week.astype(int)
    df['yw_key'] = df['iso_year'] * 100 + df['iso_week']
    
    current_yw = latest_date.isocalendar()
    current_yw_key = current_yw.year * 100 + current_yw.week
    
    weekly = df[df['yw_key'] < current_yw_key].groupby('yw_key').agg(
        High=('High', 'max'), Low=('Low', 'min'), Close=('Close', 'last'),
        Date_last=('Date', 'max')
    ).reset_index().sort_values('yw_key')
    
    if len(weekly) >= 1:
        prev_w = weekly.iloc[-1]
        levels = calc_pivot_levels(prev_w['High'], prev_w['Low'], prev_w['Close'])
        result['W1'] = {**levels, 'H': round(prev_w['High'], 2), 'L': round(prev_w['Low'], 2),
                        'C': round(prev_w['Close'], 2), 'date': prev_w['Date_last'].strftime('%Y-%m-%d')}
    
    # MN: Previous completed month
    current_ym = latest_date.strftime('%Y-%m')
    df['ym'] = df['Date'].dt.strftime('%Y-%m')
    
    monthly = df[df['ym'] < current_ym].groupby('ym').agg(
        High=('High', 'max'), Low=('Low', 'min'), Close=('Close', 'last'),
        Date_last=('Date', 'max')
    ).reset_index().sort_values('ym')
    
    if len(monthly) >= 1:
        prev_m = monthly.iloc[-1]
        levels = calc_pivot_levels(prev_m['High'], prev_m['Low'], prev_m['Close'])
        result['MN'] = {**levels, 'H': round(prev_m['High'], 2), 'L': round(prev_m['Low'], 2),
                        'C': round(prev_m['Close'], 2), 'date': prev_m['Date_last'].strftime('%Y-%m-%d')}
    
    return result, current_price

def find_confluence_zones(pivots, threshold=500):
    """
    BTC confluence threshold = 500 (wider than gold's 20 because BTC price is ~$70k).
    """
    all_levels = []
    level_names = ['R3', 'R2', 'R1', 'PP', 'S1', 'S2', 'S3']
    for tf, data in pivots.items():
        for lv in level_names:
            if lv in data:
                all_levels.append({'tf': tf, 'level': lv, 'price': data[lv]})
    
    used = set()
    clusters = []
    for i, a in enumerate(all_levels):
        if i in used:
            continue
        group = [a]
        for j, b in enumerate(all_levels):
            if j <= i or j in used:
                continue
            if a['tf'] == b['tf']:
                continue
            if abs(a['price'] - b['price']) < threshold:
                group.append(b)
                used.add(j)
        if len(group) >= 2:
            used.add(i)
            clusters.append(group)
    
    for cluster in clusters:
        cluster.sort(key=lambda x: x['price'], reverse=True)
    clusters.sort(key=lambda c: sum(x['price'] for x in c) / len(c), reverse=True)
    
    return clusters

# ── Compute pivots ──
pivots, current_price_pivot = calc_multi_tf_pivots(df)

print(f"\n{'='*55}")
print(f"Multi-TF Pivot Points")
print(f"{'='*55}")
for tf in ['D1', 'W1', 'MN']:
    if tf in pivots:
        p = pivots[tf]
        print(f"  {tf}: PP={p['PP']:.2f}  R1={p['R1']:.2f}  R2={p['R2']:.2f}  R3={p['R3']:.2f}  S1={p['S1']:.2f}  S2={p['S2']:.2f}  S3={p['S3']:.2f}  (H/L/C from {p['date']})")

confluences = find_confluence_zones(pivots, threshold=500)
if confluences:
    print(f"\n⚡ Confluence Zones (within 500 pts):")
    for cluster in confluences:
        avg = sum(x['price'] for x in cluster) / len(cluster)
        tags = " + ".join(f"{x['tf']} {x['level']}({x['price']:.2f})" for x in cluster)
        is_r = any(x['level'].startswith('R') for x in cluster)
        is_s = any(x['level'].startswith('S') for x in cluster)
        zone_type = "Resistance" if is_r and not is_s else ("Support" if is_s and not is_r else "Mixed")
        print(f"  ~{avg:.2f} — {tags} [{zone_type}]")
else:
    print(f"\n  No confluence zones found")

# ── Flatten pivot data for CSV ──
def flatten_pivots_for_csv(pivots, confluences, current_price):
    flat = {}
    for tf in ['D1', 'W1', 'MN']:
        if tf in pivots:
            for lv in ['PP', 'R1', 'R2', 'R3', 'S1', 'S2', 'S3']:
                flat[f'Pivot_{tf}_{lv}'] = pivots[tf][lv]
            flat[f'Pivot_{tf}_H'] = pivots[tf]['H']
            flat[f'Pivot_{tf}_L'] = pivots[tf]['L']
            flat[f'Pivot_{tf}_C'] = pivots[tf]['C']
            flat[f'Pivot_{tf}_Date'] = pivots[tf]['date']
        else:
            for lv in ['PP', 'R1', 'R2', 'R3', 'S1', 'S2', 'S3']:
                flat[f'Pivot_{tf}_{lv}'] = ''
            flat[f'Pivot_{tf}_H'] = ''
            flat[f'Pivot_{tf}_L'] = ''
            flat[f'Pivot_{tf}_C'] = ''
            flat[f'Pivot_{tf}_Date'] = ''
    
    conf_parts = []
    for cluster in confluences:
        avg = sum(x['price'] for x in cluster) / len(cluster)
        tags = "+".join(f"{x['tf']} {x['level']}({x['price']:.2f})" for x in cluster)
        is_r = any(x['level'].startswith('R') for x in cluster)
        is_s = any(x['level'].startswith('S') for x in cluster)
        zone_type = "Resistance" if is_r and not is_s else ("Support" if is_s and not is_r else "Mixed")
        conf_parts.append(f"~{avg:.2f}|{tags}|{zone_type}")
    flat['Confluence_Zones'] = ";".join(conf_parts) if conf_parts else 'None'
    flat['Confluence_Count'] = len(confluences)
    
    if 'D1' in pivots:
        d1 = pivots['D1']
        if current_price >= d1['R3']:
            flat['Pivot_Position'] = 'Above R3'
        elif current_price >= d1['R2']:
            flat['Pivot_Position'] = 'R2-R3'
        elif current_price >= d1['R1']:
            flat['Pivot_Position'] = 'R1-R2'
        elif current_price >= d1['PP']:
            flat['Pivot_Position'] = 'PP-R1'
        elif current_price >= d1['S1']:
            flat['Pivot_Position'] = 'S1-PP'
        elif current_price >= d1['S2']:
            flat['Pivot_Position'] = 'S2-S1'
        elif current_price >= d1['S3']:
            flat['Pivot_Position'] = 'S3-S2'
        else:
            flat['Pivot_Position'] = 'Below S3'
    else:
        flat['Pivot_Position'] = ''
    
    return flat

pivot_csv = flatten_pivots_for_csv(pivots, confluences, current_price_pivot)


# ══════════════════════════════════════════════════════
# CSV OUTPUT
# ══════════════════════════════════════════════════════

csv_row = {
    'Rank': 1,
    'Ticker': 'BTC',
    'Net_Score_Avg': round(net_avg, 2),
    'Gross_Score_Avg': round(gross_avg, 2),
    'Net_Score_BD1': round(s1['net'], 2),
    'Net_Score_BD2': round(s2['net'], 2),
    'Score_Delta': round(delta, 2),
    'Tier': momentum_tier,
    'D1_ReturnRank': round(s2['d1'], 2),
    'D2_VolumeRank': round(s2['d2'], 2),
    'D3_RSI': round(s2['d3'], 2),
    'D4_MA': round(s2['d4'], 2),
    'D5_Volatility': round(s2['d5'], 2),
    'D6_External': round(s2['d6'], 2),
    'D6_Raw': s2['d6_raw'],
    'Penalty_Scaled': round(s2['penalty_scaled'], 2),
    'WP_Return_Pct': round(s2['wp_ret'], 2),
    'WP_Volume_Pct': round(s2['wp_vol'], 2),
    'Ret_1Y_Pct': round(s2['penalties']['ret_1y'], 2),
    'Ret_6M_Pct': round(s2['penalties']['ret_6m'], 2),
    'Ret_3M_Pct': round(s2['ret_pctls']['3M']['return'], 2),
    'Ret_1M_Pct': round(s2['ret_pctls']['1M']['return'], 2),
    'Ret_1W_Pct': round(s2['ret_pctls']['1W']['return'], 2),
    'RSI_Value': round(s2['rsi'], 2),
    'MA50': round(s2['ma50'], 2) if s2['ma50'] else '',
    'MA200': round(s2['ma200'], 2) if s2['ma200'] else '',
    'Price': round(s2['price'], 2),
    'Golden_Cross': str(s2['golden_cross']),
    'Volatility_Pct': round(s2['volatility'], 2),
    'Penalty_Total': s2['penalties']['total'],
    'Penalty_Reversal': s2['penalties']['reversal'],
    'Penalty_DeathCross': s2['penalties']['death_cross'],
    'Warning_Flags': s2['penalties']['flags'] if s2['penalties']['flags'] else 'None',
    'DXY_1M_Pct': round(s2['external']['dxy_1m'], 2) if s2['external']['dxy_1m'] is not None else '',
    'VIX_Level': round(s2['external']['vix_level'], 2) if s2['external']['vix_level'] is not None else '',
    'DXY_Signal': s2['external']['dxy_signal'],
    'VIX_Signal': s2['external']['vix_signal'],
    'News_Top20': 'FALSE',
    'Base_Date_1': s1['date'].strftime('%Y-%m-%d'),
    'Base_Date_2': s2['date'].strftime('%Y-%m-%d'),
    'As_Of_Running': AS_OF,
    **pivot_csv
}

csv_df = pd.DataFrame([csv_row])
csv_fixed = os.path.join(base_dir, 'output_momentum_btc.csv')
csv_ts = os.path.join(base_dir, f'output_momentum_btc_{TS_FILE}.csv')
csv_df.to_csv(csv_fixed, index=False, encoding='utf-8')
csv_df.to_csv(csv_ts, index=False, encoding='utf-8')
print(f"\nCSV saved: {csv_fixed}")
print(f"CSV saved: {csv_ts}")


# ══════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════

wb = Workbook()

gold_fill = PatternFill('solid', fgColor='F7931A')  # Bitcoin orange
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')
blue_fill = PatternFill('solid', fgColor='BDD7EE')
purple_fill = PatternFill('solid', fgColor='E2D0F8')
gray_fill = PatternFill('solid', fgColor='D9D9D9')
dark_fill = PatternFill('solid', fgColor='333333')
header_font = Font(bold=True, size=11, color='FFFFFF')
title_font = Font(bold=True, size=14, color='333333')
big_font = Font(bold=True, size=18, color='1F4E79')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols, fill=None):
    f = fill or PatternFill('solid', fgColor='1F4E79')
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def style_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt: cell.number_format = fmt
    return cell

# Sheet 1: Summary
ws1 = wb.active
ws1.title = "BTC Momentum v3.0"
ws1.sheet_properties.tabColor = "F7931A"

ws1.merge_cells('A1:H1')
ws1['A1'] = "🪙 BTC Momentum Score v3.0 — 100-Scale Edition"
ws1['A1'].font = big_font
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:H2')
ws1['A2'] = f"Run: {AS_OF}  |  BD1: {s1['date'].strftime('%Y-%m-%d')}  |  BD2: {s2['date'].strftime('%Y-%m-%d')}"
ws1['A2'].font = Font(size=10, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')

r = 4
ws1.cell(row=r, column=1, value="Net Score Avg").font = Font(bold=True, size=12)
ws1.cell(row=r, column=2, value=round(net_avg, 2)).font = Font(bold=True, size=16, color='F7931A')
ws1.cell(row=r, column=3, value=momentum_tier).font = Font(bold=True, size=12)
r += 1
ws1.cell(row=r, column=1, value="Gross Score Avg")
ws1.cell(row=r, column=2, value=round(gross_avg, 2))
r += 1
ws1.cell(row=r, column=1, value="Score Delta")
ws1.cell(row=r, column=2, value=round(delta, 2))
ws1.cell(row=r, column=2).font = Font(color='00B050' if delta >= 0 else 'FF0000')
r += 1
ws1.cell(row=r, column=1, value="Price")
ws1.cell(row=r, column=2, value=round(s2['price'], 2)).number_format = '#,##0.0'

r += 2
headers = ['Dimension', 'BD1', 'BD2', 'Score (/100)']
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header_row(ws1, r, len(headers))

dims = [
    ('D1 Return Rank', s1['d1'], s2['d1']),
    ('D2 Volume Rank', s1['d2'], s2['d2']),
    ('D3 RSI', s1['d3'], s2['d3']),
    ('D4 MA Trend', s1['d4'], s2['d4']),
    ('D5 Volatility', s1['d5'], s2['d5']),
    ('D6 External (DXY+VIX)', s1['d6'], s2['d6']),
]
for name, v1, v2 in dims:
    r += 1
    style_cell(ws1, r, 1).value = name
    style_cell(ws1, r, 2, '0.0').value = round(v1, 1)
    style_cell(ws1, r, 3, '0.0').value = round(v2, 1)
    style_cell(ws1, r, 4, '0.0').value = round(v2, 1)

for c in range(1, 9):
    ws1.column_dimensions[get_column_letter(c)].width = 20

# Sheet 2: Methodology
ws2 = wb.create_sheet("Methodology v3.0")
ws2.sheet_properties.tabColor = "4472C4"
ws2['A1'] = "BTC Momentum Scoring v3.0 — Methodology"
ws2['A1'].font = big_font
ws2['A3'] = "Key differences from Gold:"
ws2['A3'].font = Font(bold=True, size=12)
ws2['A4'] = "D5 Volatility: thresholds adjusted for crypto (BTC 30-120% vs Gold 15-60%)"
ws2['A5'] = "D6 External: BTC is risk-ON (opposite of gold's safe-haven)"
ws2['A6'] = "  - VIX low + BTC up = bullish (calm market = risk-on)"
ws2['A7'] = "  - VIX high + BTC down = bearish (panic selling risk assets)"
ws2['A8'] = "Annualized vol uses sqrt(365) not sqrt(252) for crypto"
ws2['A9'] = "Pivot confluence threshold = 500pts (vs gold's 20pts) due to higher price"

xlsx_path = os.path.join(base_dir, 'BTC_Momentum_v3.0.xlsx')
wb.save(xlsx_path)
print(f"\nExcel saved: {xlsx_path}")
print(f"\n✅ BTC BUY score outputs generated successfully!")
